import csv
import io
import json
import os
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple, Optional

import requests
from flask import Flask, Response, flash, redirect, render_template, request, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

try:
    from google.oauth2 import service_account
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseUpload
except Exception:  # pragma: no cover
    service_account = None
    build = None
    MediaIoBaseUpload = None

BASE_DIR = Path(__file__).resolve().parent
ALLOWED_EXTENSIONS = {"xlsx", "xlsm", "pdf"}

app = Flask(__name__)
app.secret_key = os.environ.get("MATERIAL_APP_SECRET", "change-this-secret")

STATUS_OPTIONS = [
    "Chưa đặt hàng",
    "Đã đặt hàng",
]

HEADER_ALIASES = {
    # App sẽ tự nhận nhiều tên cột khác nhau. Tên được chuẩn hóa bỏ dấu, không phân biệt hoa/thường.
    "item_code": [
        "stt ma", "ma vt", "ma vat tu", "ma hang", "ma hang hoa", "ma thiet bi", "ma san pham",
        "code", "item code", "material code", "part no", "part number", "catalog no", "tag", "tag no", "ma"
    ],
    "item_name": [
        "ten vat tu", "ten hang", "ten hang hoa", "ten thiet bi", "noi dung", "hang muc", "mo ta",
        "description", "item description", "material name", "item name", "name", "material", "item", "goods", "equipment"
    ],
    "specification": [
        "quy cach", "quy cach ky thuat", "thong so", "thong so ky thuat", "dac tinh", "model",
        "spec", "specification", "technical specification", "size", "kich thuoc", "vat lieu", "material type", "type"
    ],
    "unit": ["don vi tinh", "don vi", "dvt", "unit", "uom", "u o m"],
    "quantity": ["so luong", "sl", "khoi luong", "qty", "quantity", "vol", "volume", "amount"],
    "supplier": [
        "nha cung cap", "ncc", "don vi cung cap", "supplier", "vendor", "hang sx", "hang san xuat",
        "manufacturer", "maker", "brand"
    ],
    "po_no": ["po", "po no", "po number", "so po", "ma po", "purchase order", "don hang", "so don hang", "order no"],
    "request_date": ["ngay yeu cau", "ngay de nghi", "request date", "ngay dat", "order date", "ngay mua"],
    "required_date": ["ngay can", "ngay can hang", "ngay giao", "ngay giao hang", "required date", "delivery date", "eta", "deadline"],
    "note": ["ghi chu", "ghi chu noi bo", "note", "notes", "remark", "remarks", "comment", "comments"],
}



def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    replacements = str.maketrans(
        "àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ",
        "aaaaaaaaaaaaaaaaaeeeeeeeeeeeiiiiiooooooooooooooooouuuuuuuuuuuyyyyyd",
    )
    return " ".join(text.translate(replacements).replace("_", " ").replace("/", " ").split())


def now_text() -> str:
    return datetime.now().isoformat(timespec="seconds")


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    return str(value).strip()


def file_extension(filename: str) -> str:
    return filename.rsplit(".", 1)[1].lower() if "." in filename else ""


def allowed_file(filename: str) -> bool:
    return file_extension(filename) in ALLOWED_EXTENSIONS


def is_excel_file(filename: str) -> bool:
    return file_extension(filename) in {"xlsx", "xlsm"}


def is_pdf_file(filename: str) -> bool:
    return file_extension(filename) == "pdf"


def header_cell_matches(cell: str, alias: str) -> bool:
    """So khớp tên cột theo kiểu gần đúng nhưng tránh nhầm quá rộng."""
    if not cell or not alias:
        return False
    if cell == alias:
        return True
    # Cho phép chứa nhau với alias đủ dài, ví dụ "Tên vật tư/hàng hóa" chứa "ten vat tu".
    if len(alias) >= 3 and (alias in cell or cell in alias):
        return True
    # So theo token để nhận các cột như "Material / Item Description".
    cell_tokens = set(cell.split())
    alias_tokens = set(alias.split())
    if alias_tokens and alias_tokens.issubset(cell_tokens):
        return True
    return False


def detect_header(row_values: List[Any]) -> Tuple[Dict[str, int], int]:
    normalized = [normalize_text(v) for v in row_values]
    mapping: Dict[str, int] = {}
    score = 0
    for field, aliases in HEADER_ALIASES.items():
        aliases_norm = [normalize_text(a) for a in aliases]
        for idx, cell in enumerate(normalized):
            if not cell:
                continue
            if any(header_cell_matches(cell, alias) for alias in aliases_norm):
                if field not in mapping:
                    mapping[field] = idx
                    # Các cột quan trọng cho điểm cao hơn để chọn đúng dòng tiêu đề.
                    score += 3 if field in {"item_name", "quantity", "unit"} else 1
                break
    return mapping, score


def parse_excel(path: Path) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: List[Dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            all_rows = list(sheet.iter_rows(values_only=True))
            if not all_rows:
                continue

            best_index: Optional[int] = None
            best_mapping: Dict[str, int] = {}
            best_score = 0
            # Quét 50 dòng đầu để tìm dòng tiêu đề, vì một số file có logo/ghi chú ở trên.
            for idx, row in enumerate(all_rows[:50]):
                mapping, score = detect_header(list(row))
                # Chỉ nhận dòng tiêu đề nếu có ít nhất cột tên vật tư/mô tả.
                if "item_name" not in mapping:
                    continue
                if score > best_score:
                    best_score = score
                    best_index = idx
                    best_mapping = mapping

            # Không đoán theo vị trí cứng nếu không nhận ra cột tên vật tư.
            # Làm vậy để tránh import nhầm các file không đúng nội dung vật tư.
            if best_index is None or "item_name" not in best_mapping:
                continue

            headers = [cell_to_text(v) or f"Column {i+1}" for i, v in enumerate(all_rows[best_index])]
            empty_streak = 0
            for row_number, row in enumerate(all_rows[best_index + 1 :], start=best_index + 2):
                if not any(cell_to_text(v) for v in row):
                    empty_streak += 1
                    # Nếu đã đọc qua nhiều dòng trống liên tiếp thì coi như hết bảng.
                    if empty_streak >= 15:
                        break
                    continue
                empty_streak = 0

                item: Dict[str, Any] = {"sheet_name": sheet.title, "excel_row": row_number, "extra_json": {}}
                for field, col_index in best_mapping.items():
                    item[field] = cell_to_text(row[col_index]) if col_index < len(row) else ""

                # Bỏ qua các dòng tổng cộng/ghi chú nếu không có tên vật tư thực sự.
                name_norm = normalize_text(item.get("item_name", ""))
                if not name_norm or name_norm in {"tong", "tong cong", "total", "subtotal"}:
                    continue

                for col_index, header in enumerate(headers):
                    value = cell_to_text(row[col_index]) if col_index < len(row) else ""
                    if value:
                        item["extra_json"][header] = value
                rows.append(item)
    finally:
        workbook.close()
    return rows



class SupabaseClient:
    def __init__(self) -> None:
        self.url = os.environ.get("SUPABASE_URL", "").rstrip("/")
        self.key = os.environ.get("SUPABASE_KEY", "") or os.environ.get("SUPABASE_ANON_KEY", "")
        if not self.url or not self.key:
            raise RuntimeError("Thiếu SUPABASE_URL hoặc SUPABASE_KEY trong Environment Variables.")
        self.headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        }

    def table_url(self, table: str) -> str:
        return f"{self.url}/rest/v1/{table}"

    def _request(self, method: str, table: str, **kwargs):
        response = requests.request(method, self.table_url(table), headers=self.headers, timeout=30, **kwargs)
        if response.status_code >= 400:
            raise RuntimeError(f"Supabase lỗi {response.status_code}: {response.text[:500]}")
        if response.text:
            return response.json()
        return []

    def select(self, table: str, params: Dict[str, str]) -> List[Dict[str, Any]]:
        return self._request("GET", table, params=params)

    def insert(self, table: str, payload: Any) -> List[Dict[str, Any]]:
        return self._request("POST", table, data=json.dumps(payload, ensure_ascii=False))

    def update(self, table: str, payload: Dict[str, Any], params: Dict[str, str]) -> List[Dict[str, Any]]:
        return self._request("PATCH", table, params=params, data=json.dumps(payload, ensure_ascii=False))

    def delete(self, table: str, params: Dict[str, str]) -> None:
        self._request("DELETE", table, params=params)


def sb() -> SupabaseClient:
    return SupabaseClient()


def get_drive_service():
    if service_account is None or build is None:
        raise RuntimeError("Thiếu thư viện Google API. Kiểm tra requirements.txt và redeploy.")
    raw = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    if not raw:
        raise RuntimeError("Thiếu GOOGLE_SERVICE_ACCOUNT_JSON trong Environment Variables.")
    try:
        info = json.loads(raw)
    except json.JSONDecodeError:
        # Trường hợp private_key bị dán thành 1 dòng có ký tự \\n.
        fixed = raw.replace('\\n', '\n')
        info = json.loads(fixed)
    if "private_key" in info:
        info["private_key"] = info["private_key"].replace('\\n', '\n')
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive"]
    )
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def drive_root_id() -> str:
    root_id = os.environ.get("GOOGLE_DRIVE_ROOT_FOLDER_ID", "").strip()
    if not root_id:
        raise RuntimeError("Thiếu GOOGLE_DRIVE_ROOT_FOLDER_ID trong Environment Variables.")
    return root_id


def create_drive_folder(name: str) -> Tuple[str, str]:
    service = get_drive_service()
    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [drive_root_id()],
    }
    folder = service.files().create(body=metadata, fields="id, webViewLink").execute()
    folder_id = folder["id"]
    return folder_id, folder.get("webViewLink", f"https://drive.google.com/drive/folders/{folder_id}")


def drive_mimetype(filename: str) -> str:
    ext = file_extension(filename)
    if ext in {"xlsx", "xlsm"}:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    if ext == "pdf":
        return "application/pdf"
    return "application/octet-stream"


def upload_to_drive(file_bytes: bytes, filename: str, folder_id: str) -> Tuple[str, str]:
    service = get_drive_service()
    metadata = {"name": filename, "parents": [folder_id]}
    media = MediaIoBaseUpload(
        io.BytesIO(file_bytes),
        mimetype=drive_mimetype(filename),
        resumable=False,
    )
    created = service.files().create(body=metadata, media_body=media, fields="id, webViewLink").execute()
    file_id = created["id"]
    return file_id, created.get("webViewLink", f"https://drive.google.com/file/d/{file_id}/view")



def trash_drive_item(file_id: str) -> None:
    """Đưa file/thư mục Google Drive vào thùng rác, không xóa vĩnh viễn."""
    if not file_id:
        return
    service = get_drive_service()
    service.files().update(fileId=file_id, body={"trashed": True}, fields="id, trashed").execute()


def drive_item_exists(file_id: str) -> bool:
    """Kiểm tra file/thư mục Drive còn tồn tại và chưa nằm trong thùng rác."""
    if not file_id:
        return False
    service = get_drive_service()
    try:
        item = service.files().get(fileId=file_id, fields="id, trashed").execute()
        return not bool(item.get("trashed"))
    except Exception:
        return False


@app.context_processor
def inject_globals():
    return {"STATUS_OPTIONS": STATUS_OPTIONS}


@app.route("/")
def index():
    try:
        client = sb()
        projects = client.select("projects", {"select": "*", "order": "created_at.desc"})
        materials = client.select("materials", {"select": "project_id,is_ordered"})
        counts: Dict[int, Dict[str, int]] = {}
        for m in materials:
            pid = int(m.get("project_id"))
            counts.setdefault(pid, {"total": 0, "done": 0, "issue": 0})
            counts[pid]["total"] += 1
            if m.get("is_ordered"):
                counts[pid]["done"] += 1
            else:
                counts[pid]["issue"] += 1
        for p in projects:
            c = counts.get(int(p["id"]), {"total": 0, "done": 0, "issue": 0})
            p["total_items"] = c["total"]
            p["done_items"] = c["done"]
            p["issue_items"] = c["issue"]
    except Exception as exc:
        flash(str(exc), "error")
        projects = []
    return render_template("index.html", projects=projects)


@app.route("/projects", methods=["POST"])
def create_project():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Vui lòng nhập tên dự án.", "error")
        return redirect(url_for("index"))

    # Không hiển thị mã dự án trên giao diện nữa, nhưng vẫn tự tạo mã nội bộ
    # để phù hợp với cấu trúc Supabase hiện có.
    code = f"CKQN-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    description = ""
    folder_name = name
    try:
        client = sb()
        folder_id, folder_url = create_drive_folder(folder_name)
        client.insert(
            "projects",
            {
                "name": name,
                "code": code,
                "description": description,
                "drive_folder_id": folder_id,
                "drive_folder_url": folder_url,
            },
        )
        flash("Đã tạo dự án và tự tạo thư mục trên Google Drive.", "success")
    except Exception as exc:
        flash(f"Không tạo được dự án: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/project/<int:project_id>")
def project_detail(project_id: int):
    status = request.args.get("status", "")
    keyword = request.args.get("q", "").strip().lower()
    try:
        client = sb()
        projects = client.select("projects", {"select": "*", "id": f"eq.{project_id}", "limit": "1"})
        if not projects:
            flash("Không tìm thấy dự án.", "error")
            return redirect(url_for("index"))
        project = projects[0]
        params = {"select": "*", "project_id": f"eq.{project_id}", "order": "id.desc"}
        if status == "Đã đặt hàng":
            params["is_ordered"] = "eq.true"
        elif status == "Chưa đặt hàng":
            params["is_ordered"] = "eq.false"
        materials = client.select("materials", params)
        # Chuẩn hóa tên field cho template cũ.
        for m in materials:
            m["item_code"] = m.get("material_code") or ""
            m["item_name"] = m.get("material_name") or ""
            m["status"] = "Đã đặt hàng" if m.get("is_ordered") else "Chưa đặt hàng"
        if keyword:
            def hit(m: Dict[str, Any]) -> bool:
                hay = " ".join(str(m.get(k, "")) for k in [
                    "material_name", "material_code", "specification", "po_no", "supplier", "note"
                ]).lower()
                return keyword in hay
            materials = [m for m in materials if hit(m)]
        files = client.select("excel_files", {"select": "*", "project_id": f"eq.{project_id}", "order": "uploaded_at.desc"})
        for f in files:
            f["original_name"] = f.get("file_name") or ""
            f["row_count"] = f.get("row_count") or 0
        ordered_count = sum(1 for m in materials if m.get("is_ordered"))
        not_ordered_count = len(materials) - ordered_count
        stats = [
            {"status": "Chưa đặt hàng", "count": not_ordered_count},
            {"status": "Đã đặt hàng", "count": ordered_count},
        ]
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("index"))
    return render_template(
        "project.html", project=project, materials=materials, files=files,
        stats=stats, status=status, keyword=keyword
    )


@app.route("/project/<int:project_id>/upload", methods=["POST"])
def upload_excel(project_id: int):
    uploaded = request.files.get("order_file") or request.files.get("excel_file")
    if not uploaded or not uploaded.filename:
        flash("Vui lòng chọn file Excel hoặc PDF.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not allowed_file(uploaded.filename):
        flash("Chỉ hỗ trợ file .xlsx, .xlsm hoặc .pdf.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    try:
        client = sb()
        projects = client.select("projects", {"select": "*", "id": f"eq.{project_id}", "limit": "1"})
        if not projects:
            flash("Không tìm thấy dự án.", "error")
            return redirect(url_for("index"))
        project = projects[0]
        folder_id = project.get("drive_folder_id")
        if not folder_id:
            raise RuntimeError("Dự án này chưa có drive_folder_id. Hãy tạo lại dự án hoặc cập nhật folder ID.")

        file_bytes = uploaded.read()
        original_name = secure_filename(uploaded.filename) or "upload"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        drive_name = f"{timestamp}_{original_name}"

        rows: List[Dict[str, Any]] = []
        if is_excel_file(original_name):
            with tempfile.NamedTemporaryFile(suffix=Path(original_name).suffix, delete=False) as tmp:
                tmp.write(file_bytes)
                tmp_path = Path(tmp.name)
            try:
                rows = parse_excel(tmp_path)
            finally:
                tmp_path.unlink(missing_ok=True)

        drive_file_id, drive_file_url = upload_to_drive(file_bytes, drive_name, folder_id)
        file_payload = {
            "project_id": project_id,
            "file_name": uploaded.filename,
            "drive_file_id": drive_file_id,
            "drive_file_url": drive_file_url,
            "row_count": len(rows),
        }
        if is_pdf_file(original_name):
            file_payload["file_type"] = "pdf"
        else:
            file_payload["file_type"] = "excel"
        inserted_file = client.insert("excel_files", file_payload)[0]
        file_id = inserted_file["id"]

        if rows:
            payload = []
            for item in rows:
                payload.append(
                    {
                        "project_id": project_id,
                        "excel_file_id": file_id,
                        "sheet_name": item.get("sheet_name"),
                        "excel_row": item.get("excel_row"),
                        "material_code": item.get("item_code", ""),
                        "material_name": item.get("item_name", ""),
                        "specification": item.get("specification", ""),
                        "unit": item.get("unit", ""),
                        "quantity": item.get("quantity", ""),
                        "supplier": item.get("supplier", ""),
                        "po_no": item.get("po_no", ""),
                        "request_date": item.get("request_date", ""),
                        "required_date": item.get("required_date", ""),
                        "note": item.get("note", ""),
                        "is_ordered": False,
                        "extra_json": item.get("extra_json", {}),
                        "updated_at": now_text(),
                    }
                )
            # Supabase REST giới hạn payload quá lớn, chia nhỏ mỗi 500 dòng.
            for start in range(0, len(payload), 500):
                client.insert("materials", payload[start : start + 500])
            flash(f"Đã upload file Excel lên Google Drive và import {len(rows)} dòng vật tư.", "success")
        elif is_pdf_file(original_name):
            flash("Đã upload file PDF lên Google Drive. PDF hiện chỉ lưu hồ sơ gốc, chưa tự import vật tư vào bảng.", "success")
        else:
            flash("Đã upload file lên Google Drive nhưng không có dòng vật tư nào được import.", "success")
    except Exception as exc:
        flash(f"Không upload/import được file: {exc}", "error")
    return redirect(url_for("project_detail", project_id=project_id))


@app.route("/material/<int:material_id>/update", methods=["POST"])
def update_material(material_id: int):
    status = request.form.get("status", "Chưa đặt hàng")
    is_ordered = status == "Đã đặt hàng"
    fields = {
        "is_ordered": is_ordered,
        "location": request.form.get("location", "").strip(),
        "responsible": request.form.get("responsible", "").strip(),
        "actual_date": request.form.get("actual_date", "").strip(),
        "note": request.form.get("note", "").strip(),
        "updated_at": now_text(),
    }
    try:
        client = sb()
        rows = client.select("materials", {"select": "project_id", "id": f"eq.{material_id}", "limit": "1"})
        if not rows:
            flash("Không tìm thấy vật tư.", "error")
            return redirect(url_for("index"))
        project_id = rows[0]["project_id"]
        client.update("materials", fields, {"id": f"eq.{material_id}"})
        flash("Đã cập nhật tình trạng vật tư.", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("index"))


@app.route("/project/<int:project_id>/export.csv")
def export_csv(project_id: int):
    try:
        client = sb()
        projects = client.select("projects", {"select": "*", "id": f"eq.{project_id}", "limit": "1"})
        if not projects:
            return redirect(url_for("index"))
        project = projects[0]
        materials = client.select("materials", {"select": "*", "project_id": f"eq.{project_id}", "order": "id.asc"})
    except Exception as exc:
        flash(str(exc), "error")
        return redirect(url_for("index"))
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Mã vật tư", "Tên vật tư", "Quy cách", "Đơn vị", "Số lượng", "Nhà cung cấp", "PO",
        "Ngày yêu cầu", "Ngày cần", "Tình trạng", "Vị trí", "Người phụ trách", "Ngày thực tế", "Ghi chú"
    ])
    for m in materials:
        writer.writerow([
            m.get("material_code", ""), m.get("material_name", ""), m.get("specification", ""),
            m.get("unit", ""), m.get("quantity", ""), m.get("supplier", ""), m.get("po_no", ""),
            m.get("request_date", ""), m.get("required_date", ""),
            "Đã đặt hàng" if m.get("is_ordered") else "Chưa đặt hàng",
            m.get("location", ""), m.get("responsible", ""), m.get("actual_date", ""), m.get("note", "")
        ])
    filename = f"{secure_filename(project.get('name') or 'materials')}_materials.csv"
    return Response(
        output.getvalue().encode("utf-8-sig"),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/project/<int:project_id>/delete", methods=["POST"])
def delete_project(project_id: int):
    try:
        client = sb()
        projects = client.select("projects", {"select": "id,drive_folder_id", "id": f"eq.{project_id}", "limit": "1"})
        if not projects:
            flash("Không tìm thấy dự án cần xóa.", "error")
            return redirect(url_for("index"))
        folder_id = projects[0].get("drive_folder_id")
        if folder_id:
            trash_drive_item(folder_id)
        # Xóa rõ các bảng con trước để không phụ thuộc cascade.
        client.delete("materials", {"project_id": f"eq.{project_id}"})
        client.delete("excel_files", {"project_id": f"eq.{project_id}"})
        client.delete("projects", {"id": f"eq.{project_id}"})
        flash("Đã xóa dự án trên app và đưa thư mục Google Drive vào thùng rác.", "success")
    except Exception as exc:
        flash(f"Không xóa được dự án: {exc}", "error")
    return redirect(url_for("index"))


@app.route("/sync-drive", methods=["POST"])
def sync_drive():
    """Đồng bộ thủ công: nếu thư mục Drive đã bị xóa/đưa vào thùng rác thì xóa dự án trên app."""
    try:
        client = sb()
        projects = client.select("projects", {"select": "id,name,drive_folder_id", "order": "created_at.desc"})
        removed = 0
        for project in projects:
            project_id = project.get("id")
            folder_id = project.get("drive_folder_id")
            if not drive_item_exists(folder_id):
                client.delete("materials", {"project_id": f"eq.{project_id}"})
                client.delete("excel_files", {"project_id": f"eq.{project_id}"})
                client.delete("projects", {"id": f"eq.{project_id}"})
                removed += 1
        if removed:
            flash(f"Đã đồng bộ Google Drive và xóa {removed} dự án không còn thư mục Drive.", "success")
        else:
            flash("Đã đồng bộ Google Drive. Không có dự án nào cần xóa.", "success")
    except Exception as exc:
        flash(f"Không đồng bộ được Google Drive: {exc}", "error")
    return redirect(url_for("index"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug)
